import numpy as np
from openpyxl import load_workbook
from networkx import from_numpy_matrix, draw, set_node_attributes, draw_networkx_labels, connected_components, \
    compose_all  # networkx is a library used for graph operations and plotting
import matplotlib.pyplot as plt
from plotly.figure_factory._county_choropleth import create_choropleth  # this function creates a choropleth plot


class CountyMap:
    def __init__(self):
        # initialize variables and dictionaries of information
        self.COUNTY_LIST = dict()
        self.COUNTY_FIPS = dict()
        self.COUNTY_LOCATIONS = dict()
        self.COUNTY_TECHNIQUES = dict()
        self.adjacencyMatrix = None
        self.NODE_COLORS_BY_TECHNIQUE = dict()
        self.TECHNIQUES = dict()

    def load_counties(self, source_xlsx):  # the data in source_xlsx needs to be formatted so that the first column
        # lists out all the counties that are present in the map area
        workbook = load_workbook(filename=source_xlsx)
        sheet = workbook.active
        for row in range(2, sheet.max_row):  # upon loading the workbook, we iterate from the 2nd row forwards
            if sheet.cell(row, 1).value is None:  # stop iteration if we encounter an empty cell
                break
            # we store the "index" of each county in the dictionary COUNTY_LIST and its location coordinates in
            # COUNTY_LOCATIONS, as supplied in the sheet
            self.COUNTY_LIST[sheet.cell(row, 1).value] = row - 2
            self.COUNTY_FIPS[row - 2] = sheet.cell(row, 9).value
            self.COUNTY_TECHNIQUES[row - 2] = 0
            self.COUNTY_LOCATIONS[row - 2] = (int(sheet.cell(row, 3).value),
                                              int(sheet.cell(row, 4).value))
        self.adjacencyMatrix = np.zeros((len(self.COUNTY_LIST), len(self.COUNTY_LIST)))  # create the adjacency matrix
        # of size n x n for n counties

    def load_connections(self, source_xlsx):  # the data in source_xlsx needs to be formatted so that the third and
        # fourth columns list out pairs of counties that have an edge between them
        # in our graph
        workbook = load_workbook(filename=source_xlsx)
        sheet = workbook.active
        for row in range(2, sheet.max_row):  # upon loading the workbook, we iterate from the 2nd row forwards
            if sheet.cell(row, 6).value is None or sheet.cell(row, 7).value is None:  # stop iteration if cell is empty
                break
            firstCounty = sheet.cell(row, 6).value
            secondCounty = sheet.cell(row, 7).value
            # once we have our two counties determined, we find their index from our dictionary and add a link between
            # the two counties in the adjacency matrix
            self.adjacencyMatrix[self.COUNTY_LIST[firstCounty]][self.COUNTY_LIST[secondCounty]] = 1

    def load_techniques(self, techniques_dictionary):
        self.TECHNIQUES = techniques_dictionary

    def load_node_colors(self, node_colors_by_technique_dictionary):
        self.NODE_COLORS_BY_TECHNIQUE = node_colors_by_technique_dictionary

    def assign_technique_to_county(self, county, technique):
        if technique not in self.TECHNIQUES:
            raise ValueError("Technique is not present in list.")
        self.COUNTY_TECHNIQUES[county] = technique

    def connected_components_by_technique(self, technique):
        techniqueGraph = from_numpy_matrix(self.adjacencyMatrix)  # create networkx graph object
        for county, technique_used in self.COUNTY_TECHNIQUES.items():
            if technique != technique_used:
                techniqueGraph.remove_node(county)  # remove all counties that don't use the given irrigation technique

        return connected_components(techniqueGraph)  # return the number of connected components

    def draw_graph(self):
        graph = from_numpy_matrix(self.adjacencyMatrix)  # create networkx graph object
        set_node_attributes(graph, self.COUNTY_LOCATIONS, 'coord')  # set attributes for coordinates of each county on
                                                                    # the graph
        # the rest of the code plots the graph in a pyplot figure
        plt.figure(figsize=(10, 10))
        plt.margins(x=0.2)
        draw(graph, pos=self.COUNTY_LOCATIONS, node_size=150)
        draw_networkx_labels(graph, pos={k: (v[0] + 0.4, v[1] + 0.25) for k, v in self.COUNTY_LOCATIONS.items()},
                             labels={v: k for k, v in self.COUNTY_LIST.items()}, font_size=11)

        plt.show()

    def draw_graph_by_technique(self, technique):
        techniqueGraph = from_numpy_matrix(self.adjacencyMatrix)  # create networkx graph object
        includedCounties = []
        for county, technique_used in self.COUNTY_TECHNIQUES.items():  # remove counties that do not use the technique
            if technique != technique_used:
                techniqueGraph.remove_node(county)
            else:
                includedCounties.append(county)

        techniqueGraphNodeList = {k: v for k, v in self.COUNTY_LIST.items() if v in includedCounties}
        techniqueGraphNodeLocations = {k: v for k, v in self.COUNTY_LOCATIONS.items() if k in includedCounties}

        # the rest of the code is similar to that of self.draw_graph()
        set_node_attributes(techniqueGraph, techniqueGraphNodeLocations, 'coord')
        plt.figure(figsize=(10, 10))
        plt.margins(x=0.2)
        draw(techniqueGraph, pos=techniqueGraphNodeLocations, node_size=150,
             node_color=self.NODE_COLORS_BY_TECHNIQUE[technique])
        draw_networkx_labels(techniqueGraph, pos={k: (v[0] + 0.4, v[1] + 0.25)
                                                  for k, v in techniqueGraphNodeLocations.items()},
                             labels={v: k for k, v in techniqueGraphNodeList.items()}, font_size=11)

        plt.show()

    def draw_graph_for_all_techniques(self):
        techniqueGraphs = []
        plt.figure(figsize=(10, 10))
        plt.margins(x=0.2)
        for technique, name in self.TECHNIQUES.items():  # we create networkx graph objects
            techniqueGraph = from_numpy_matrix(self.adjacencyMatrix)
            includedCounties = []
            for county, technique_used in self.COUNTY_TECHNIQUES.items():
                if technique != technique_used:
                    techniqueGraph.remove_node(county)
                else:
                    includedCounties.append(county)

            techniqueGraphs.append(techniqueGraph)
            # we plot every graph in the same way as self.draw_graph() and self.draw_graph_by_technique() and then
            # overlay them on top of one another to get our final graph for all techniques
            techniqueGraphNodeList = {k: v for k, v in self.COUNTY_LIST.items() if v in includedCounties}
            techniqueGraphNodeLocations = {k: v for k, v in self.COUNTY_LOCATIONS.items() if k in includedCounties}

            set_node_attributes(techniqueGraph, techniqueGraphNodeLocations, 'coord')
            set_node_attributes(techniqueGraph, self.NODE_COLORS_BY_TECHNIQUE[technique], 'color')

            draw(techniqueGraph, pos=techniqueGraphNodeLocations, node_size=150,
                 node_color=self.NODE_COLORS_BY_TECHNIQUE[technique])
            draw_networkx_labels(techniqueGraph, pos={k: (v[0] + 0.4, v[1] + 0.25)
                                                      for k, v in techniqueGraphNodeLocations.items()},
                                 labels={v: k for k, v in techniqueGraphNodeList.items()}, font_size=11)

        plt.show()

    def county_choropleth_by_technique(self):
        techniques = [v for k, v in self.COUNTY_TECHNIQUES.items()]
        fips = [v for k, v in self.COUNTY_FIPS.items()]
        color_scale = [v for k, v in self.NODE_COLORS_BY_TECHNIQUE.items()]
        # the choropleth graphs are generated with the scope of Texas and use the color scale described in
        # self.NODE_COLORS_BY_TECHNIQUE
        fig = create_choropleth(fips=fips, values=techniques, scope=['Texas'], plot_bgcolor='rgb(229,229,229)',
                                paper_bgcolor='rgb(229,229,229)',
                                legend_title='Irrigation Technique by County',
                                state_outline={'color': 'rgb(0, 0, 0)', 'width': 2},
                                county_outline={'color': 'rgb(255, 255, 255)', 'width': 0.5},
                                colorscale=color_scale)

        fig.show()
