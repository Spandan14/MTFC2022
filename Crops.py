from openpyxl import load_workbook

# the following classes contain data for each crop type and also functions for calculating water usage by crop

CURRENT_AVERAGE_EFFICIENCY = 0.6285


class Corn:
    def __init__(self, owning_county, acres_planted, bushel_yield):
        self.owning_county = owning_county
        self.acres_planted = acres_planted
        self.bushel_yield = bushel_yield
        self.LBS_PER_BUSHEL = 56
        self.GALLONS_PER_LB = 73.4398 * CURRENT_AVERAGE_EFFICIENCY

    def water_usage(self):
        return self.bushel_yield * self.LBS_PER_BUSHEL * self.GALLONS_PER_LB


class Sorghum:
    def __init__(self, owning_county, acres_planted, bushel_yield):
        self.owning_county = owning_county
        self.acres_planted = acres_planted
        self.bushel_yield = bushel_yield
        self.LBS_PER_BUSHEL = 50
        self.GALLONS_PER_LB = 143.1813 * CURRENT_AVERAGE_EFFICIENCY

    def water_usage(self):
        return self.bushel_yield * self.LBS_PER_BUSHEL * self.GALLONS_PER_LB


class Wheat:
    def __init__(self, owning_county, acres_planted, bushel_yield):
        self.owning_county = owning_county
        self.acres_planted = acres_planted
        self.bushel_yield = bushel_yield
        self.LBS_PER_BUSHEL = 60
        self.GALLONS_PER_LB = 268.7951 * CURRENT_AVERAGE_EFFICIENCY

    def water_usage(self):
        return self.bushel_yield * self.LBS_PER_BUSHEL * self.GALLONS_PER_LB


class Cotton:
    def __init__(self, owning_county, acres_planted, lbs_per_acre_yield):
        self.owning_county = owning_county
        self.acres_planted = acres_planted
        self.lbs_per_acre_yield = lbs_per_acre_yield
        self.GALLONS_PER_LB = 653.0333 * CURRENT_AVERAGE_EFFICIENCY

    def water_usage(self):
        return self.acres_planted * self.lbs_per_acre_yield * self.GALLONS_PER_LB


class Peanuts:
    def __init__(self, owning_county, acres_planted, lbs_yield):
        self.owning_county = owning_county
        self.acres_planted = acres_planted
        self.lbs_yield = lbs_yield
        self.GALLONS_PER_LB = 127.8593 * CURRENT_AVERAGE_EFFICIENCY

    def water_usage(self):
        return self.lbs_yield * self.GALLONS_PER_LB


def crop_class_loader(source_xlsx):  # loads crop classes into a list by county from a source xlsx file
    county_crops = []
    workbook = load_workbook(filename=source_xlsx)  # load workbook
    sheet = workbook.active
    for row in range(2, sheet.max_row):  # iterate from 2nd row of sheet
        if sheet.cell(row, 1).value is None:  # stop iteration if cell is blank
            break

        current_county_crops = []
        # the following lines of code simply check the 2nd, 4th, 6th, 8th, and 10th columns for data on the 5 crops
        # we consider; if data is found, then a class instance is made for the crop with appropriate parameters and
        # added to the county currently under consideration
        if sheet.cell(row, 2).value is not None:
            current_county_crops.append(Corn(owning_county=row - 2,
                                             acres_planted=sheet.cell(row, 2).value,
                                             bushel_yield=sheet.cell(row, 3).value))
        if sheet.cell(row, 4).value is not None:
            current_county_crops.append(Wheat(owning_county=row - 2,
                                              acres_planted=sheet.cell(row, 4).value,
                                              bushel_yield=sheet.cell(row, 5).value))
        if sheet.cell(row, 6).value is not None:
            current_county_crops.append(Sorghum(owning_county=row - 2,
                                                acres_planted=sheet.cell(row, 6).value,
                                                bushel_yield=sheet.cell(row, 7).value))
        if sheet.cell(row, 8).value is not None:
            current_county_crops.append(Cotton(owning_county=row - 2,
                                               acres_planted=sheet.cell(row, 8).value,
                                               lbs_per_acre_yield=sheet.cell(row, 9).value))
        if sheet.cell(row, 10).value is not None:
            current_county_crops.append(Peanuts(owning_county=row - 2,
                                                acres_planted=sheet.cell(row, 10).value,
                                                lbs_yield=sheet.cell(row, 11).value))

        county_crops.append(current_county_crops)

    return county_crops