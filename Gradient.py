from openpyxl import load_workbook


def gradient_loader(source_xlsx):  # this function simply loads county gradients from a source xlsx file
    county_gradients = []
    workbook = load_workbook(filename=source_xlsx)  # load workbook
    sheet = workbook.active
    for row in range(2, sheet.max_row):  # iterate from 2nd row of sheet
        if sheet.cell(row, 1).value is None:  # stop iteration if cell is blank
            break

        county_gradients.append(sheet.cell(row, 2).value)  # insert value into gradient list

    return county_gradients
